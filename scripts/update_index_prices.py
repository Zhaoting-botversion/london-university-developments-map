"""Sync bedroom pricing data from the enriched Excel workbook into index.html.

Usage from the vault root:
  python "💼 工作专项/London University Developments Map/scripts/update_index_prices.py"
  python "💼 工作专项/London University Developments Map/scripts/update_index_prices.py" --write

The script replaces only the JavaScript `PRICE_LOOKUP` block. It does not touch
the university/project coordinate data in `UNIVERSITIES_STUB`.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_DIR / "University community developments - enriched prices.xlsx"
INDEX_HTML = PROJECT_DIR / "index.html"

SHEET_TO_KEY = {
    "UCL (Bloomsbury)": "UCL",
    "UCL (East)": "UCLE",
    "KCL (Strand Waterloo) & LSE": "KCL",
    "帝国理工 IC 主校区": "IC",
    "Chelsea College of Arts": "CCA",
    "中央圣马丁": "CSM",
    "伦敦传媒学院 LCC": "LCC",
}

PRICE_COLUMNS = {
    "studio": "开间价格",
    "one": "一居室价格",
    "two": "二居室价格",
    "three": "三居室价格",
    "fourPlus": "四居及以上价格",
    "penthouse": "Penthouse价格",
}

META_COLUMNS = {
    "reservation": "预定金",
    "payment": "付款条件",
    "source": "价单来源",
}


def clean_header(value: object) -> str:
    return str(value or "").strip()


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def js_stringify(value: object) -> str:
    text = json.dumps(value, ensure_ascii=False, indent=2)
    return text.replace("\u2028", "\\u2028").replace("\u2029", "\\u2029")


def printable_path(path: Path) -> str:
    return str(path).encode("ascii", "backslashreplace").decode("ascii")


def build_price_lookup(workbook_path: Path) -> dict[str, dict[str, dict[str, object]]]:
    wb = load_workbook(workbook_path, data_only=True)
    lookup: dict[str, dict[str, dict[str, object]]] = {}

    for ws in wb.worksheets:
        sheet_name = ws.title.strip()
        uni_key = SHEET_TO_KEY.get(sheet_name)
        if not uni_key:
            continue

        headers = {
            clean_header(cell.value): idx + 1
            for idx, cell in enumerate(ws[1])
            if clean_header(cell.value)
        }
        name_col = headers.get("项目名") or headers.get("项目名称")
        if not name_col:
            raise ValueError(f"Cannot find project name column in sheet: {ws.title}")

        projects: dict[str, dict[str, object]] = {}
        for row_idx in range(2, ws.max_row + 1):
            project_name = clean_cell(ws.cell(row=row_idx, column=name_col).value)
            if not project_name:
                continue

            prices = {
                key: clean_cell(ws.cell(row=row_idx, column=headers[column]).value)
                if column in headers
                else ""
                for key, column in PRICE_COLUMNS.items()
            }
            details = {
                "prices": prices,
                **{
                    key: clean_cell(ws.cell(row=row_idx, column=headers[column]).value)
                    if column in headers
                    else ""
                    for key, column in META_COLUMNS.items()
                },
            }
            projects[project_name] = details

        lookup[uni_key] = projects

    missing = sorted(set(SHEET_TO_KEY.values()) - set(lookup))
    if missing:
        raise ValueError(f"Missing expected university sheets for keys: {', '.join(missing)}")

    return lookup


def replace_price_lookup(index_text: str, lookup: dict[str, object]) -> str:
    pattern = re.compile(
        r"const PRICE_LOOKUP = \{[\s\S]*?\n\};(?=\n\nfunction csvToRows)",
        re.MULTILINE,
    )
    replacement = (
        "const PRICE_LOOKUP = "
        + js_stringify(lookup)
        + ";"
    )
    new_text, count = pattern.subn(replacement, index_text, count=1)
    if count != 1:
        raise ValueError("Could not locate exactly one PRICE_LOOKUP block in index.html")
    return new_text


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--index", type=Path, default=INDEX_HTML)
    parser.add_argument("--write", action="store_true", help="Write changes to index.html")
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create an index.html backup when writing",
    )
    args = parser.parse_args()

    lookup = build_price_lookup(args.workbook)
    index_text = args.index.read_text(encoding="utf-8")
    new_text = replace_price_lookup(index_text, lookup)

    if new_text == index_text:
        print("PRICE_LOOKUP is already up to date.")
        return 0

    if not args.write:
        print("PRICE_LOOKUP would be updated. Re-run with --write to apply.")
        return 1

    if not args.no_backup:
        backup_path = args.index.with_suffix(".backup-before-price-sync.html")
        copy2(args.index, backup_path)
        print(f"Backup created: {printable_path(backup_path)}")

    args.index.write_text(new_text, encoding="utf-8", newline="\n")
    print(f"Updated: {printable_path(args.index)}")
    print(f"Universities synced: {', '.join(lookup.keys())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
